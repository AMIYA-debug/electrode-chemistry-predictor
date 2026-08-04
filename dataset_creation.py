from __future__ import annotations

import argparse
import logging
import math
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook as openpyxl_load_workbook


LOGGER = logging.getLogger(__name__)
MISSING_FORMULATION_VALUES = {"", "-", "none", "nan", "na", "n/a"}
USELESS_LABELS = {"balance", "fe balance", "remarks", "remark", "total", "totals"}


@dataclass(frozen=True)
class TableRegion:
    """Coordinates and metadata for a detected table in a raw worksheet."""

    sheet: str
    header_row: int
    first_data_row: int
    last_data_row: int
    label_column: int
    value_columns: tuple[int, ...]
    reason: str


def load_workbook(path: Path) -> tuple[pd.ExcelFile, dict[str, pd.DataFrame]]:
    """Load all worksheets without assigning Excel cells as DataFrame headers."""
    if not path.is_file():
        raise FileNotFoundError(f"Workbook was not found: {path.resolve()}")
    try:
        excel = pd.ExcelFile(path)
        sheets = {
            name: pd.read_excel(excel, sheet_name=name, header=None, dtype=object)
            for name in excel.sheet_names
        }
    except Exception as exc:  
        raise RuntimeError(f"Could not read Excel workbook '{path}': {exc}") from exc
    return excel, sheets


def _is_blank(value: Any) -> bool:
    return pd.isna(value) or (isinstance(value, str) and not value.strip())


def _normalise(value: Any) -> str:
    return "" if _is_blank(value) else str(value).strip().casefold()


def _sample_key(value: Any) -> str | None:
    """Return a stable sample identifier, accepting Excel's 1 and 1.0 alike."""
    if _is_blank(value) or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if isinstance(value, float) and math.isnan(value):
            return None
        return str(int(value)) if float(value).is_integer() else str(value).strip()
    text = str(value).strip()
    if re.fullmatch(r"\d+\.0+", text):
        return text.split(".", 1)[0]
    return text or None


def _looks_like_number(value: Any) -> bool:
    return not pd.isna(pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0])


def _non_empty_rows(frame: pd.DataFrame) -> list[int]:
    return [int(index) for index, row in frame.iterrows() if row.notna().any()]


def _non_empty_regions(frame: pd.DataFrame) -> list[tuple[int, int]]:
    """Group adjacent non-empty rows into human-readable worksheet regions."""
    rows = _non_empty_rows(frame)
    if not rows:
        return []
    regions: list[tuple[int, int]] = []
    start = end = rows[0]
    for row in rows[1:]:
        if row == end + 1:
            end = row
        else:
            regions.append((start, end))
            start = end = row
    regions.append((start, end))
    return regions


def inspect_workbook(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    """Log workbook diagnostics requested for transparent, auditable inference."""
    LOGGER.info("Workbook: %s", path.resolve())
    LOGGER.info("Sheet names: %s", list(sheets))
    try:
        openpyxl_book = openpyxl_load_workbook(path, read_only=False, data_only=True)
    except Exception as exc:
        raise RuntimeError(f"Could not inspect merged cells in '{path}': {exc}") from exc

    for sheet_name, frame in sheets.items():
        merged = [str(rng) for rng in openpyxl_book[sheet_name].merged_cells.ranges]
        hidden_rows = [row_number for row_number, dimension in openpyxl_book[sheet_name].row_dimensions.items() if dimension.hidden]
        LOGGER.info("Sheet '%s': dimensions=%s, non-empty rows=%s", sheet_name, frame.shape, _non_empty_rows(frame))
        LOGGER.info("Sheet '%s': non-empty regions (zero-based rows)=%s", sheet_name, _non_empty_regions(frame))
        LOGGER.info("Sheet '%s': merged cells=%s", sheet_name, merged or "none")
        LOGGER.info("Sheet '%s': hidden Excel rows=%s", sheet_name, hidden_rows or "none")
        LOGGER.info("Sheet '%s': first 20 rows:\n%s", sheet_name, frame.head(20).to_string())


def _contiguous_formulation_end(frame: pd.DataFrame, start: int, label_col: int, sample_cols: list[int]) -> int:
    """Find the last ingredient row, stopping at the first blank separator."""
    end = start - 1
    for row_index in range(start, len(frame)):
        label = frame.iat[row_index, label_col]
        values = frame.iloc[row_index, sample_cols]
        if _is_blank(label) and values.isna().all():
            break

        if not _is_blank(label) and values.notna().any():
            end = row_index
        elif end >= start:
            break
    return end


def find_formulation_section(sheets: dict[str, pd.DataFrame]) -> TableRegion:
    candidates: list[TableRegion] = []
    for sheet_name, frame in sheets.items():
        for row_index, row in frame.iterrows():
            ingredient_columns = [
                col for col, value in row.items()
                if any(token in _normalise(value) for token in ("ingredient", "formulation", "flux"))
            ]
            for label_col in ingredient_columns:
                sample_cols = [col for col in range(label_col + 1, frame.shape[1]) if _sample_key(row[col]) is not None]
                if len(sample_cols) < 2:
                    continue
                end = _contiguous_formulation_end(frame, row_index + 1, label_col, sample_cols)
                ingredient_rows = max(0, end - row_index)
                if ingredient_rows >= 2:
                    candidates.append(TableRegion(
                        sheet_name, int(row_index), int(row_index + 1), end, int(label_col), tuple(sample_cols),
                        f"header contains '{row[label_col]}' and has {len(sample_cols)} sample columns and {ingredient_rows} ingredient rows",
                    ))
    if not candidates:
        raise ValueError("No formulation section found. Expected a header containing Ingredient, Formulation, or Flux followed by sample IDs.")
    selected = max(candidates, key=lambda item: (item.last_data_row - item.first_data_row + 1) * len(item.value_columns))
    LOGGER.info("Selected formulation section: sheet='%s', rows %d-%d, label column %d: %s",
                selected.sheet, selected.header_row, selected.last_data_row, selected.label_column, selected.reason)
    if len(candidates) > 1:
        LOGGER.warning("Other formulation candidates were found and rejected: %s", candidates)
    return selected


def find_chemistry_section(sheets: dict[str, pd.DataFrame]) -> TableRegion:

    candidates: list[TableRegion] = []
    for sheet_name, frame in sheets.items():
        for row_index, row in frame.iterrows():
            sample_columns = [col for col, value in row.items() if _normalise(value) in {"sample", "sample id", "sample no", "electrode", "electrode id"}]
            for label_col in sample_columns:
                value_cols = [col for col in range(label_col + 1, frame.shape[1]) if not _is_blank(row[col])]
                data_rows = [idx for idx in range(row_index + 1, len(frame)) if _sample_key(frame.iat[idx, label_col]) is not None]
                if len(value_cols) < 1 or len(data_rows) < 2:
                    continue

                measurement_rows = [
                    idx for idx in range(min(data_rows), len(frame))
                    if frame.iloc[idx, value_cols].map(lambda value: not _is_blank(value)).any()
                ]
                numeric_columns = sum(any(_looks_like_number(frame.iat[idx, col]) for idx in data_rows) for col in value_cols)
                if numeric_columns:
                    candidates.append(TableRegion(
                        sheet_name, int(row_index), min(data_rows), max(measurement_rows), int(label_col), tuple(value_cols),
                        f"header contains '{row[label_col]}' and {numeric_columns} columns contain measured numeric values for {len(data_rows)} named samples",
                    ))
    if not candidates:
        raise ValueError("No chemistry section found. Expected a header named Sample (or Sample ID / Electrode) followed by chemistry columns.")
    selected = max(candidates, key=lambda item: len(item.value_columns) * (item.last_data_row - item.first_data_row + 1))
    LOGGER.info("Selected chemistry section: sheet='%s', header row %d, sample column %d: %s",
                selected.sheet, selected.header_row, selected.label_column, selected.reason)
    return selected


def extract_formulation(frame: pd.DataFrame, region: TableRegion) -> pd.DataFrame:
    """Transpose ingredient rows into one feature row per electrode sample."""
    names = frame.iloc[region.first_data_row:region.last_data_row + 1, region.label_column].tolist()
    data = frame.iloc[region.first_data_row:region.last_data_row + 1, list(region.value_columns)].copy()
    sample_ids = [_sample_key(frame.iat[region.header_row, column]) for column in region.value_columns]
    if any(sample is None for sample in sample_ids):
        raise ValueError("Formulation header includes an empty sample identifier.")
    if len(set(sample_ids)) != len(sample_ids):
        raise ValueError(f"Duplicate formulation sample IDs: {sample_ids}")
    if len(set(names)) != len(names):
        raise ValueError("Duplicate ingredient names found; names must be unique for ML feature columns.")
    result = data.T
    result.columns = names   
    result.index = pd.Index(sample_ids, name="Sample")
    return result


def extract_chemistry(frame: pd.DataFrame, region: TableRegion) -> pd.DataFrame:

    rows: list[int] = []
    sample_ids: list[str] = []
    active_sample: str | None = None
    for row_index in range(region.first_data_row, region.last_data_row + 1):
        explicit_sample = _sample_key(frame.iat[row_index, region.label_column])
        if explicit_sample is not None:
            active_sample = explicit_sample
        has_measurement = frame.iloc[row_index, list(region.value_columns)].map(lambda value: not _is_blank(value)).any()
        if active_sample is not None and has_measurement:
            rows.append(row_index)
            sample_ids.append(active_sample)
    if not rows:
        raise ValueError("Chemistry section contains no rows with a sample ID and measured values.")
    names = [frame.iat[region.header_row, column] for column in region.value_columns]
    result = frame.iloc[rows, list(region.value_columns)].copy()
    result.columns = names 
    result.index = pd.Index(sample_ids, name="Sample")
    LOGGER.info("Extracted %d chemistry measurements mapped to %d sample IDs (merged sample cells are forward-filled).",
                len(result), result.index.nunique())
    return result


def _parse_diameter_mm(value: Any) -> float | None:
    """Extract a millimetre diameter from values such as 'coating dia. 4.8'."""
    if _is_blank(value):
        return None

    match = re.search(r"(?<!\d)(\d+(?:\.\d+)?)\s*(?:mm)?\b", str(value), flags=re.IGNORECASE)
    return float(match.group(1)) if match else None


def filter_48mm_samples(chemistry: pd.DataFrame) -> pd.DataFrame:

    diameter_columns = [
        column for column in chemistry.columns
        if "diameter" in _normalise(column) or "dia" in _normalise(column)
    ]
    if not diameter_columns:
        raise ValueError("No diameter metadata column found in chemistry. Expected a header containing 'Dia' or 'Diameter' to enforce 4.8 mm-only data.")
    if len(diameter_columns) > 1:
        raise ValueError(f"Ambiguous diameter metadata columns: {diameter_columns}")
    diameter_column = diameter_columns[0]
    diameters = chemistry[diameter_column].map(_parse_diameter_mm)
    unparseable = chemistry.index[diameters.isna()].tolist()
    if unparseable:
        raise ValueError(f"Could not read electrode diameter for chemistry samples: {unparseable}")
    selected = chemistry.loc[diameters.eq(4.8)].copy()
    if selected.empty:
        available = sorted(set(diameters.tolist()))
        raise ValueError(f"No 4.8 mm chemistry samples found. Detected diameters: {available}")
    duplicated = selected.index[selected.index.duplicated()].unique().tolist()
    if duplicated:
        raise ValueError(f"Multiple 4.8 mm chemistry rows found for the same sample IDs: {duplicated}")
    LOGGER.info("Diameter column '%s' selected %d of %d measurement rows at 4.8 mm for sample IDs: %s. "
                "Ignored non-4.8 measurement rows associated with: %s",
                diameter_column, len(selected), len(chemistry), selected.index.tolist(),
                chemistry.index[~diameters.eq(4.8)].tolist())
    return selected.drop(columns=diameter_column)


def clean_formulation(formulation: pd.DataFrame) -> pd.DataFrame:

    def normalise_missing(value: Any) -> Any:
        return 0 if _normalise(value) in MISSING_FORMULATION_VALUES else value

    cleaned = formulation.map(normalise_missing)
    numeric = cleaned.apply(pd.to_numeric, errors="coerce")
    bad = numeric.isna() & ~cleaned.isna()
    if bad.any().any():
        locations = [
            (str(cleaned.index[row]), str(cleaned.columns[column]), cleaned.iat[row, column])
            for row, column in zip(*bad.to_numpy().nonzero())
        ]
        raise ValueError(f"Non-numeric formulation values cannot be converted to float: {locations[:10]}")
    return numeric.fillna(0.0).astype(float)


def clean_chemistry(chemistry: pd.DataFrame) -> pd.DataFrame:
    """Retain columns whose observed values are exclusively numeric; never zero-fill targets."""
    retained: dict[Any, pd.Series] = {}
    excluded: list[str] = []
    for name in chemistry.columns:
        if _normalise(name) in USELESS_LABELS:
            excluded.append(str(name))
            continue
        raw = chemistry[name]
        numeric = pd.to_numeric(raw, errors="coerce")
        observed = raw.map(lambda value: not _is_blank(value))

        if observed.any() and numeric[observed].notna().all():
            retained[name] = numeric.astype(float)
        else:
            excluded.append(str(name))
    if not retained:
        raise ValueError("Chemistry section has no wholly numeric target columns.")
    LOGGER.info("Chemistry targets retained: %s; excluded non-numeric/useless columns: %s", list(retained), excluded)
    return pd.DataFrame(retained, index=chemistry.index)


def merge_dataset(features: pd.DataFrame, targets: pd.DataFrame) -> pd.DataFrame:

    if list(features.index) != list(targets.index):
        raise ValueError("Sample order or membership differs between formulation and chemistry. "
                         f"Formulation={list(features.index)}; chemistry={list(targets.index)}")
    return pd.concat([features, targets], axis=1)


def validate_dataset(dataset: pd.DataFrame, features: pd.DataFrame, targets: pd.DataFrame) -> None:
    if len(features) != len(targets):
        raise ValueError(f"Sample count mismatch: {len(features)} formulation rows vs {len(targets)} chemistry rows.")
    if features.index.duplicated().any() or targets.index.duplicated().any():
        raise ValueError("Duplicated sample IDs detected.")
    duplicates = dataset.columns[dataset.columns.duplicated()].tolist()
    if duplicates:
        raise ValueError(f"Duplicate output column names detected: {duplicates}")
    if targets.isna().any().any():
        missing = targets.isna().sum()
        missing = missing[missing.gt(0)].to_dict()
        raise ValueError(f"Missing chemistry target values are not allowed: {missing}")
    if not all(pd.api.types.is_float_dtype(dtype) for dtype in dataset.dtypes):
        raise TypeError(f"Every ML column must be float; found: {dataset.dtypes.to_dict()}")
    LOGGER.info("Validation passed: %d matching samples, %d unique numeric columns.", len(dataset), len(dataset.columns))


def save_dataset(dataset: pd.DataFrame, csv_path: Path, xlsx_path: Path) -> None:

    output = dataset.reset_index()
    try:
        output.to_csv(csv_path, index=False)
        output.to_excel(xlsx_path, index=False)
    except Exception as exc:
        raise RuntimeError(f"Could not save dataset outputs: {exc}") from exc
    LOGGER.info("Saved CSV: %s", csv_path.resolve())
    LOGGER.info("Saved Excel: %s", xlsx_path.resolve())


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert welding formulation Excel data into an ML dataset.")
    parser.add_argument("workbook", nargs="?", type=Path, default=Path("Flux Formulation & chemistry 1.xlsx"))
    parser.add_argument("--output-dir", type=Path, default=Path("."))
    parser.add_argument("--log-level", default="INFO", choices=("DEBUG", "INFO", "WARNING", "ERROR"))
    args = parser.parse_args()
    logging.basicConfig(level=getattr(logging, args.log_level), format="%(levelname)s %(message)s")
    args.output_dir.mkdir(parents=True, exist_ok=True)

    _, sheets = load_workbook(args.workbook)
    inspect_workbook(args.workbook, sheets)
    formulation_region = find_formulation_section(sheets)
    chemistry_region = find_chemistry_section(sheets)
    all_features = clean_formulation(extract_formulation(sheets[formulation_region.sheet], formulation_region))
    chemistry_with_metadata = extract_chemistry(sheets[chemistry_region.sheet], chemistry_region)
    chemistry_48mm = filter_48mm_samples(chemistry_with_metadata)
    try:
        features = all_features.loc[chemistry_48mm.index]
    except KeyError as exc:
        missing = [sample for sample in chemistry_48mm.index if sample not in all_features.index]
        raise ValueError(f"4.8 mm chemistry samples have no matching formulation: {missing}") from exc
    targets = clean_chemistry(chemistry_48mm)
    dataset = merge_dataset(features, targets)
    validate_dataset(dataset, features, targets)
    save_dataset(dataset, args.output_dir / "weld_dataset.csv", args.output_dir / "weld_dataset.xlsx")

    LOGGER.info("Dataset shape (samples, ML columns): %s", dataset.shape)
    LOGGER.info("Number of 4.8 mm samples: %d", len(dataset))
    LOGGER.info("Feature count: %d", features.shape[1])
    LOGGER.info("Target count: %d", targets.shape[1])
    LOGGER.info("Feature names: %s", list(features.columns))
    LOGGER.info("Target names: %s", list(targets.columns))
    LOGGER.info("Missing values by column: %s", dataset.isna().sum().to_dict())
    LOGGER.info("Data types: %s", dataset.dtypes.astype(str).to_dict())
    LOGGER.info("First five rows:\n%s", dataset.head().to_string())


if __name__ == "__main__":
    main()
